#!/usr/bin/python
# -*- coding: utf-8 -*-
#
# Copyright (c) 2025, Marco Noce <nce.marco@gmail.com>
# GNU General Public License v3.0+ (see LICENSES/GPL-3.0-or-later.txt or https://www.gnu.org/licenses/gpl-3.0.txt)
# SPDX-License-Identifier: GPL-3.0-or-later

from __future__ import absolute_import, division, print_function
__metaclass__ = type

DOCUMENTATION = r'''
---
module: open_excel
short_description: Read and update Excel (.xlsx) files using openpyxl
version_added: "1.1.9"
author:
  - "Marco Noce (@NomakCooper)"
requirements:
  - openpyxl
description:
  - This module reads from or writes to Excel (.xlsx) files using the openpyxl Python library.
  - It supports reading the entire workbook or a single worksheet, optionally limited to a given cell range.
  - For updates, you can overwrite cells, append new rows, or insert rows. You can also apply custom cell styles.
  - The original Excel file is not overwritten unless you set O(dest) to the same path as O(src).
  - This module supports only .xlsx or .xlsm files.
options:
  src:
    description:
      - Path to the source Excel file.
    required: true
    type: str
  dest:
    description:
      - Destination file path for updated Excel content.
      - If omitted, defaults to appending V(_updated.xlsx) to the O(src) filename.
    required: false
    type: str
  op:
    description: >
      The operation to perform on the Excel file. Options:
      V(r) - Read-only. Returns the content from the specified sheet or all sheets.
      V(w) - Write. Overwrites specified cells with new values.
      V(a) - Append. Creates one new row at the end of the sheet, writing each item in O(updates_matrix) to that row.
      V(i) - Insert. Inserts a new row above the row specified in the first item of O(updates_matrix) and writes the updates.
    required: true
    type: str
    choices: ['r', 'w', 'a', 'i']
  sheet_name:
    description:
      - Name of the worksheet to operate on.
      - For O(op=r), if omitted, all sheets are read.
      - For O(op=w), O(op=a), or O(op=i), this parameter is required.
    required: false
    type: str
  index_by_name:
    description:
      - For read operations, if true, uses the first row as dictionary keys.
        Otherwise, keys are in the format V(col_<n>).
    required: false
    default: true
    type: bool
  read_range:
    description:
      - Dictionary specifying the cell range to read.
      - Can include V(start_row), V(end_row), V(start_col), and V(end_col).
      - If omitted or partially specified, defaults to the entire used range.
    required: false
    type: dict
    default: {}
  updates_matrix:
    description: >
      A list of dictionaries describing the cells to update. Each dictionary can include:
      V(cell_row) - The row to update (ignored in append mode).
      V(cell_col) - The column to update.
      V(cell_value) - The value to write.
    required: false
    type: list
    elements: dict
    default: []
  cell_style:
    description: >
      A dictionary specifying optional style attributes for updated cells. Possible keys include:
      V(fontColor) - Hex RGB code for the font color.
      V(bgColor) - Hex RGB code for the cell background color.
      V(bold) - Boolean to set bold font.
      V(italic) - Boolean to set italic font.
      V(underline) - Boolean to set underline; if true, uses single underline.
    required: false
    type: dict
    default: {}
notes:
  - This module requires the openpyxl Python library to be installed.
  - Only .xlsx or .xlsm files are supported.
'''

EXAMPLES = r'''
- name: Read Excel workbook
  nomakcooper.collection.open_excel:
    src: "/tmp/sample.xlsx"
    op: "r"
    index_by_name: true
  register: result
- debug:
    var: result

- name: Overwrite specific cells
  nomakcooper.collection.open_excel:
    src: "/tmp/sample.xlsx"
    dest: "/tmp/sample_updated.xlsx"
    op: "w"
    sheet_name: "Sheet1"
    updates_matrix:
      - cell_row: 2
        cell_col: 1
        cell_value: "New Value in row2 col1"
      - cell_row: 3
        cell_col: 2
        cell_value: "Another Value"
    cell_style:
      fontColor: "FF0000"
      bgColor: "FFFF00"
      bold: true

- name: Append new row
  nomakcooper.collection.open_excel:
    src: "/tmp/sample.xlsx"
    dest: "/tmp/sample_updated.xlsx"
    op: "a"
    sheet_name: "Sheet1"
    updates_matrix:
      - cell_col: 1
        cell_value: "Hostname"
      - cell_col: 2
        cell_value: "MyHost"
    cell_style:
      bgColor: "DDEBF7"
      bold: true

- name: Insert a new row above row 5
  nomakcooper.collection.open_excel:
    src: "/tmp/sample.xlsx"
    dest: "/tmp/sample_updated.xlsx"
    op: "i"
    sheet_name: "Sheet1"
    updates_matrix:
      - cell_row: 5
        cell_col: 1
        cell_value: "Inserted"
      - cell_row: 5
        cell_col: 2
        cell_value: "Row"
    cell_style:
      italic: true
'''

RETURN = r'''
changed:
  description: >
    Indicates whether the Excel file was modified. For read operations, this is still returned as true,
    even though no changes were made to the file.
  returned: always
  type: bool
  sample: true
result:
  description: >
    For read operations, returns a dictionary keyed by sheet name. Each key maps to a list of dictionaries,
    where each dictionary represents a row with cell values. The keys for each row are determined by the
    header row (if index_by_name is true) or default to "col_<n>".
    For write, append, and insert operations, returns an empty dictionary upon a successful update.
  returned: always
  type: dict
  sample: {
      "Sheet1": [
          {"Name": "Alice", "Age": 30},
          {"Name": "Bob", "Age": 25}
      ]
  }
'''

from ansible.module_utils.basic import AnsibleModule
from ansible.module_utils.basic import missing_required_lib
import traceback

OPENPYXL_IMPORT_ERROR = ""
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
except ImportError:
    HAS_OPENPYXL = False
    OPENPYXL_IMPORT_ERROR = traceback.format_exc()
    openpyxl = None
else:
    HAS_OPENPYXL = True


def apply_cell_style(cell, cell_style):
    if not cell_style:
        return

    font_kwargs = {}
    if 'fontColor' in cell_style and cell_style['fontColor']:
        font_kwargs['color'] = cell_style['fontColor']
    if 'bold' in cell_style:
        font_kwargs['bold'] = cell_style['bold']
    if 'italic' in cell_style:
        font_kwargs['italic'] = cell_style['italic']
    if 'underline' in cell_style:
        font_kwargs['underline'] = 'single' if cell_style['underline'] else None

    if font_kwargs:
        cell.font = Font(**font_kwargs)

    if 'bgColor' in cell_style and cell_style['bgColor']:
        cell.fill = PatternFill(fill_type="solid", fgColor=cell_style['bgColor'])


def read_excel(module, src, index_by_name, read_range, sheet_name):
    try:
        wb = openpyxl.load_workbook(src, data_only=True)
    except Exception as e:
        module.fail_json(msg="Error loading workbook: %s" % str(e))

    result = {}
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            module.fail_json(msg="Sheet name '%s' not found in workbook." % sheet_name)
        sheets_to_read = [sheet_name]
    else:
        sheets_to_read = wb.sheetnames

    for name in sheets_to_read:
        sheet = wb[name]
        start_row = int(read_range.get('start_row', 1))
        start_col = int(read_range.get('start_col', 1))
        end_row = int(read_range.get('end_row', sheet.max_row))
        end_col = int(read_range.get('end_col', sheet.max_column))

        sheet_data = []
        if index_by_name:
            keys = []
            for col in range(start_col, end_col + 1):
                val = sheet.cell(row=start_row, column=col).value
                keys.append(str(val) if val is not None else "col_%d" % col)
            data_start = start_row + 1
        else:
            keys = ["col_%d" % col for col in range(start_col, end_col + 1)]
            data_start = start_row

        for row in range(data_start, end_row + 1):
            row_dict = {}
            for idx, col in enumerate(range(start_col, end_col + 1)):
                row_dict[keys[idx]] = sheet.cell(row=row, column=col).value
            sheet_data.append(row_dict)

        result[name] = sheet_data

    return result


def update_excel(module, src, dest, updates_matrix, cell_style, sheet_name, op):
    try:
        wb = openpyxl.load_workbook(src)
    except Exception as e:
        module.fail_json(msg="Error loading workbook: %s" % str(e))

    if sheet_name not in wb.sheetnames:
        module.fail_json(msg="Sheet name '%s' not found in workbook." % sheet_name)
    sheet = wb[sheet_name]

    if op == 'a':
        new_row = sheet.max_row + 1
        for update in updates_matrix:
            col = int(update.get('cell_col', 0))
            cell = sheet.cell(row=new_row, column=col)
            cell.value = update.get('cell_value', None)
            apply_cell_style(cell, cell_style)
    elif op == 'i':
        if not updates_matrix:
            module.fail_json(msg="No updates_matrix provided for insert operation.")
        insert_row = int(updates_matrix[0].get('cell_row', 0))
        if insert_row < 1:
            module.fail_json(msg="Invalid cell_row for insert operation: %d" % insert_row)
        sheet.insert_rows(idx=insert_row, amount=1)
        for update in updates_matrix:
            col = int(update.get('cell_col', 0))
            cell = sheet.cell(row=insert_row, column=col)
            cell.value = update.get('cell_value', None)
            apply_cell_style(cell, cell_style)
    elif op == 'w':
        for update in updates_matrix:
            row = int(update.get('cell_row', 0))
            col = int(update.get('cell_col', 0))
            if row < 1 or col < 1:
                module.fail_json(msg="Invalid cell_row or cell_col in write operation.")
            cell = sheet.cell(row=row, column=col)
            cell.value = update.get('cell_value', None)
            apply_cell_style(cell, cell_style)
    else:
        module.fail_json(msg="Invalid operation: %s" % op)

    if not dest:
        dest = src.rsplit('.', 1)[0] + '_updated.xlsx'

    try:
        wb.save(dest)
    except Exception as e:
        module.fail_json(msg="Error saving workbook: %s" % str(e))

    return {}


def main():
    module = AnsibleModule(
        argument_spec=dict(
            src=dict(required=True, type='str'),
            dest=dict(required=False, type='str'),
            op=dict(required=True, type='str', choices=['r', 'w', 'a', 'i']),
            sheet_name=dict(required=False, type='str'),
            index_by_name=dict(required=False, type='bool', default=True),
            read_range=dict(required=False, type='dict', default={}),
            updates_matrix=dict(required=False, type='list', elements='dict', default=[]),
            cell_style=dict(required=False, type='dict', default={}),
        ),
        supports_check_mode=False
    )

    if not HAS_OPENPYXL:
        module.fail_json(
            msg=missing_required_lib('openpyxl'),
            exception=OPENPYXL_IMPORT_ERROR
        )

    if openpyxl is None:
        module.fail_json(msg="openpyxl module is required. Please install it using pip.")

    src = module.params['src']
    dest = module.params.get('dest')
    op = module.params['op']
    sheet_name = module.params.get('sheet_name')
    index_by_name = module.params['index_by_name']
    read_range = module.params.get('read_range') or {}
    updates_matrix = module.params.get('updates_matrix') or []
    cell_style = module.params.get('cell_style') or {}

    if op == 'r':
        result = read_excel(module, src, index_by_name, read_range, sheet_name)
    else:
        if not sheet_name:
            module.fail_json(msg="Parameter sheet_name is required for write operations ('w', 'a', 'i').")
        result = update_excel(module, src, dest, updates_matrix, cell_style, sheet_name, op)

    module.exit_json(changed=True, result=result)


if __name__ == '__main__':
    main()
